import json
import requests
import time
from openpyxl import Workbook
from .config import Config

class XiechengAgent:
    def __init__(self):
        self.config = Config()
        self.headers = self.config.headers
        self.request_url = self.config.request_url
        # 存储有效省-市ID组合的列表（最终写入Excel）
        self.valid_location_list = []

    def get_hotels(self, city_id:int, province_id:int, check_in_date, check_out_date, country_id=1):
        """
        核心方法：请求酒店接口，返回是否存在有效城市（通过城市名称判断）
        返回值：(是否有效, 省份名称, 城市名称)
        """
        # 设置城市和日期（日期用固定值，仅为校验ID有效性）
        self.config.set_location(city_id, province_id, country_id)
        self.config.set_dates(check_in_date, check_out_date)
        
        try:
            # 发送请求（添加1秒延迟，避免触发接口反爬）
            time.sleep(1)
            response = requests.post(
                url=self.request_url,
                headers=self.headers,
                json=self.config.request_body,
                timeout=15
            )

            if response.status_code == 200:
                try:
                    hotel_data = response.json()
                    # 提取酒店列表（无论是否有酒店，先获取城市名称）
                    hotel_list = hotel_data.get("data", {}).get("hotelList", [])
                    
                    # 从第一家酒店的信息中提取“省份名称”和“城市名称”（判断ID是否有效）
                    if hotel_list:
                        hotel_info = hotel_list[0].get("hotelInfo", {})
                        position_info = hotel_info.get("positionInfo", {})
                        province_name = position_info.get("provinceName", "未知省份")  # 省份名称
                        city_name = position_info.get("cityName", "未知城市")        # 城市名称
                        # 有效判断：城市名称不是“未知城市”，且与ID匹配（排除无效ID）
                        if city_name != "未知城市" and str(city_id) != "-1":
                            print(f"✅ 有效组合：province_id={province_id}({province_name}) | city_id={city_id}({city_name})")
                            return (True, province_name, city_name)
                except json.JSONDecodeError:
                    pass
            # 状态码非200或解析失败，视为无效
            return (False, "未知省份", "未知城市")
            
        except Exception as e:
            # 网络错误等异常，视为无效
            print(f"❌ 请求异常（province_id={province_id}, city_id={city_id}）：{str(e)}")
            return (False, "未知省份", "未知城市")

    def traverse_location_ids(self, province_id_range, city_id_range, check_in_date, check_out_date):
        """
        遍历省份ID和城市ID的指定范围，筛选有效组合
        :param province_id_range: 省份ID范围（如 range(1, 30) 表示遍历1-29）
        :param city_id_range: 城市ID范围（如 range(1, 50) 表示遍历1-49）
        """
        print(f"开始遍历省-市ID，省份范围：{province_id_range.start}-{province_id_range.stop-1}，城市范围：{city_id_range.start}-{city_id_range.stop-1}")
        
        # 第一个for循环：遍历省份ID
        for province_id in province_id_range:
            print(f"\n===== 正在遍历省份ID：{province_id} =====")
            
            # 第二个for循环：遍历该省份下的城市ID
            for city_id in city_id_range:
                # 调用get_hotels判断当前ID组合是否有效
                is_valid, province_name, city_name = self.get_hotels(
                    city_id=city_id,
                    province_id=province_id,
                    check_in_date=check_in_date,
                    check_out_date=check_out_date
                )
                
                # 若有效，添加到列表（去重：避免同一城市被多次添加）
                if is_valid:
                    valid_item = {
                        "province_id": province_id,
                        "province_name": province_name,
                        "city_id": city_id,
                        "city_name": city_name
                    }
                    # 去重判断：检查是否已存在相同的省ID+市ID组合
                    if not any(
                        item["province_id"] == province_id and item["city_id"] == city_id 
                        for item in self.valid_location_list
                    ):
                        self.valid_location_list.append(valid_item)

    def save_to_excel(self, excel_path="valid_province_city_ids.xlsx"):
        """
        将有效省-市ID组合写入Excel文件
        :param excel_path: 输出Excel路径（默认当前目录）
        """
        if not self.valid_location_list:
            print("⚠️  无有效省-市ID组合，无需生成Excel")
            return

        # 创建Excel工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "有效省-市ID组合"

        # 写入Excel表头
        headers = ["省份ID", "省份名称", "城市ID", "城市名称"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # 写入有效数据（从第2行开始）
        for row, item in enumerate(self.valid_location_list, 2):
            ws.cell(row=row, column=1, value=item["province_id"])
            ws.cell(row=row, column=2, value=item["province_name"])
            ws.cell(row=row, column=3, value=item["city_id"])
            ws.cell(row=row, column=4, value=item["city_name"])

        # 保存Excel
        wb.save(excel_path)
        print(f"\n✅ 有效省-市ID组合已保存到：{excel_path}")
        print(f"📊 共获取到 {len(self.valid_location_list)} 组有效组合")


if __name__ == "__main__":
    # 1. 初始化代理类
    agent = XiechengAgent()

    # 2. 定义要遍历的ID范围（根据携程编码规律，合理设置范围，避免无效遍历）
    # 省份ID范围：携程省份ID通常在1-30之间（覆盖全国各省/直辖市）
    province_range = range(1, 33)  
    # 城市ID范围：每个省份的城市ID通常在1-50之间（可根据需求扩大）
    city_range = range(1, 600)      

    # 3. 固定日期（仅用于校验ID有效性，无需真实入住日期）
    check_in = "20250917"
    check_out = "20250918"

    # 4. 执行遍历：循环省份和城市ID，筛选有效组合
    agent.traverse_location_ids(
        province_id_range=province_range,
        city_id_range=city_range,
        check_in_date=check_in,
        check_out_date=check_out
    )

    # 5. 将有效组合写入Excel
    agent.save_to_excel(excel_path="携程省市区ID对照表.xlsx")